import asyncio
import math
import os
import re
import time
from io import BytesIO
from typing import Set

import aiohttp
import jieba
import pandas as pd
import requests
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor  # 关键导入

# 读取brand.txt解析为列表
# 获取当前文件夹路径
current_dir = os.path.dirname(os.path.abspath(__file__))

brand_filepath = current_dir + "/../../data/brand.txt"
BRAND_DICTIONARY = set(line.strip() for line in open(brand_filepath, encoding="utf-8"))
print("已加载品牌列表：", BRAND_DICTIONARY)
dict_filepath = current_dir + "/../../data/jieba_dict.txt"

jieba.load_userdict(dict_filepath)
print("已加载自定义分词词典：", jieba.get_dict_file())


def load_excel(file_path: str) -> pd.DataFrame:
    """
    读取Excel文件，验证必要列并处理空值
    要求文件必须包含"商品ID"和"商品名称"列（可修改required_cols适配实际列名）
    """
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        required_cols = ['商品ID', '商品名称']
        # 检查必要列是否存在
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise ValueError(f"缺少必要列：{', '.join(missing_cols)}，需包含{required_cols}")
        # 去除商品ID或名称为空的无效行
        df = df.dropna(subset=required_cols).reset_index(drop=True)
        # 强制转为字符串类型，避免数字ID/名称拼接出错（如科学计数法、格式丢失）
        df['商品ID'] = df['商品ID'].astype(str).str.strip()
        df['商品名称'] = df['商品名称'].astype(str).str.strip()
        return df
    except Exception as e:
        print(f"读取文件{file_path}失败：{str(e)}")
        raise


async def save_excel_async(df: pd.DataFrame, file_path: str):
    df.to_excel(file_path, index=False, engine='openpyxl')
    print(f"\n下载完成！结果已保存到：{file_path}")


def clean_text(text) -> str:
    """清理文本，只保留中文、英文、数字和常用符号"""
    if not isinstance(text, str):
        return ""
    text = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text.lower()


def chinese_tokenize_sync(text: str) -> set:
    """jieba 同步分词"""
    if not text or text.isspace():
        return set()
    return set(jieba.cut(text, cut_all=False))


async def chinese_tokenize(text: str) -> set:
    """异步分词包装"""
    return await asyncio.to_thread(chinese_tokenize_sync, text)


def extract_full_spec(text) -> str:
    """"提取规格信息"""
    if not isinstance(text, str):
        return ""
    units = ["盒", "箱", "件", "个", "袋", "包", "瓶", "罐", "卷", "片", "只",
             "ml", "l", "g", "kg", "mm", "cm", "m"]
    pack_units = ["盒", "箱", "件", "个", "袋", "包", "瓶", "罐", "卷", "片", "只", "条", "支"]
    pattern = re.compile(
        rf'((?:\d+(?:\.\d+)?\s?(?:{"|".join(units)}|{"|".join(pack_units)}))'
        rf'(?:[*_×xX/\\-]?\s?\d*(?:\.\d+)?\s?(?:{"|".join(units)}|{"|".join(pack_units)}))*)',
        flags=re.IGNORECASE
    )
    matches = pattern.findall(text)
    if not matches:
        return ""
    specs = sorted(matches, key=len, reverse=True)
    return specs[0].strip()


def extract_brand(text_to_search: str, brand_dictionary: Set[str]) -> str:
    """提取品牌信息"""
    if not text_to_search or text_to_search.isspace() or not brand_dictionary:
        return ""
    for brand_info in brand_dictionary:
        if brand_info in text_to_search:
            return brand_info
    return ""


def pandas_str_to_series(s) -> pd.Series:
    """字符串转为Series"""
    # 判断是否已经是 Series
    if not isinstance(s, str):
        return s
    # s为None或""或nan
    if s is None or s == "" or (isinstance(s, float) and math.isnan(s)):
        return None

    inner = s[s.find("(") + 1: s.rfind(")")]

    pattern = re.compile(r"(\w+)=('[^']*'|[^,]*)")
    data = {k: v.strip("'") if v.strip() != "nan" else None for k, v in pattern.findall(inner)}

    # 3️⃣ 转为 DataFrame
    df = pd.DataFrame([data])
    return df.iloc[0]


def download_image(url):
    """根据 URL 下载图片，返回 BytesIO 对象"""
    try:
        r = requests.get(url, timeout=5)
        if r.status_code == 200:
            return BytesIO(r.content)
    except Exception as e:
        print(e)
        return None
    return None


async def fetch_image(session, url, semaphore):
    """单个图片下载任务，受信号量控制并发数"""
    if not url or str(url) in ["nan", "None", ""]:
        return url, None

    async with semaphore:  # 限制并发数量
        try:
            async with session.get(url, timeout=10) as response:
                if response.status == 200:
                    content = await response.read()
                    return url, content
        except Exception as e:
            # print(f"下载失败: {url} - {e}") # 调试时可开启
            pass
    return url, None


async def download_all_images(urls, concurrency=20):
    """并发下载所有唯一的URL"""
    tasks = []
    # 去重，避免重复下载
    unique_urls = set(urls)

    # 限制并发数为 20（根据网络情况调整）
    semaphore = asyncio.Semaphore(concurrency)

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }

    async with aiohttp.ClientSession(headers=headers) as session:
        for url in unique_urls:
            tasks.append(fetch_image(session, url, semaphore))

        # 等待所有任务完成
        results = await asyncio.gather(*tasks)
        return dict(results)


import nest_asyncio  # 新增

nest_asyncio.apply()


def pic_download(df: pd.DataFrame, output_path):
    print("正在分析图片链接...")

    # 定义包含图片的列名
    img_cols = ['origin_url', 'llm_image_url', 'top1_image_url', 'top2_image_url', 'top3_image_url', 'top4_image_url',
                'top5_image_url']

    # 1. 提取所有需要下载的 URL
    all_urls = []
    for col in img_cols:
        if col in df.columns:
            all_urls.extend(df[col].dropna().astype(str).tolist())

    print(f"检测到 {len(all_urls)} 个链接，开始异步并发下载...")
    start_time = time.time()

    # 2. 执行异步下载
    # 兼容 Windows 的事件循环策略
    try:
        loop = asyncio.get_event_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    img_data_map = loop.run_until_complete(download_all_images(all_urls))

    print(f"下载完成，耗时: {time.time() - start_time:.2f}秒。开始写入 Excel...")

    # 3. 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.append(df.columns.tolist())  # 写入表头

    # 预先设置好列宽（提升性能，避免循环中重复设置）
    # 图片列宽设为 18 (约对应 120px 宽度)，其他列默认
    for idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        if col_name in img_cols:
            ws.column_dimensions[col_letter].width = 18
        else:
            ws.column_dimensions[col_letter].width = 15

    # 4. 遍历数据写入
    row_idx = 2
    for _, row in df.iterrows():
        # 设置行高 (设为 90 以容纳缩略图)
        ws.row_dimensions[row_idx].height = 90

        col_idx = 1
        for col_name in df.columns:
            value = row[col_name]
            cell_val = str(value) if value is not None else ""

            # 判断是否为图片列
            if col_name in img_cols:
                # 尝试从缓存中获取图片数据
                img_bytes = img_data_map.get(cell_val)

                if img_bytes:
                    try:
                        # 处理图片
                        pil_img = PILImage.open(BytesIO(img_bytes))

                        # 转换格式防止报错 (如 WebP -> RGB)
                        if pil_img.mode != 'RGB':
                            pil_img = pil_img.convert('RGB')

                        # 缩放图片 (保持比例)
                        pil_img.thumbnail((120, 120))

                        # 保存为字节流供 openpyxl 使用
                        output = BytesIO()
                        pil_img.save(output, format="PNG")
                        output.seek(0)

                        # 创建 Openpyxl 图片对象
                        img = Image(output)

                        _col_start = col_idx - 1
                        _row_start = row_idx - 1

                        # 右下角锚点 (即下一个单元格的左上角)
                        _col_end = col_idx
                        _row_end = row_idx

                        marker_from = AnchorMarker(col=_col_start, colOff=0, row=_row_start, rowOff=0)
                        marker_to = AnchorMarker(col=_col_end, colOff=0, row=_row_end, rowOff=0)

                        # editAs='twoCell' 对应 Excel 中的 "Move and size with cells"
                        img.anchor = TwoCellAnchor(editAs='twoCell', _from=marker_from, to=marker_to)

                        # 添加图片 (此时不需要指定 cell_pos，因为 anchor 已经包含了位置信息)
                        ws.add_image(img)
                        # 如果需要，也可以保留文本值作为备份（可选，会覆盖在图片下或显示在公式栏）
                        # ws.cell(row=row_idx, column=col_idx, value=cell_val)

                    except Exception as e:
                        print(f"图片处理出错 [行{row_idx}]: {e}")
                        ws.cell(row=row_idx, column=col_idx, value=cell_val)
                else:
                    # 下载失败或无图片，写入链接文本
                    ws.cell(row=row_idx, column=col_idx, value=cell_val)
            else:
                # 普通列
                ws.cell(row=row_idx, column=col_idx, value=cell_val)

            col_idx += 1
        row_idx += 1

    wb.save(output_path)
    print(f"数据写入完成！文件保存至：{output_path}")
